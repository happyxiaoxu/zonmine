from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

font = Font(
                name='Calibri',
                size=9,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000'
            )

# test = PatternFill(patternType='solid',
#                                         fill_type='solid', 
#                                         fgColor=Color('C4C4C4'))

search_column_headers = ['keyword']
scout_column_headers = [
                    'search_keyword', 
                    'search_rank', 
                    'asin',
                    'sponsored',
                    'title', 
                    'brand', 
                    'price', 
                    'bread_crumbs', 
                    'bsr', 
                    'bsr_category', 
                    'review_count', 
                    'review_score', 
                    'five_star_percentage', 
                    'four_star_percentage', 
                    'three_star_percentage', 
                    'two_star_percentage', 
                    'one_star_percentage', 
                    'sold_by', 
                    'sellers_count', 
                    'url', 
                    'in_stock', 
                    'item_weight', 
                    'primary_image', 
                    'image1',
                    'image2',
                    'image3',
                    'image4',
                    'image5',
                    'image6',
                    'image7',
                    # 'image_8', 
                    # 'product_images', 
                    'description', 
                    'is_fba', 
                    # 'dimensions', 
                    'is_amz', 
                    'item_model_number', 
                    'manufacturer', 
                    # 'weight', 
                    'is_addon', 
                    'is_fbm', 
                    'is_prime', 
                    'item_dimensions_length', 
                    'item_dimensions_width', 
                    'item_dimensions_thickness', 
                    'feature1',
                    'feature2',
                    'feature3',
                    'feature4',
                    'feature5',
                    'cpc',
                    'monthly_search_volume',
                    'competition',
                ]



def init_excel(bot_name):
    wb = Workbook()
    sheet = wb.active
    if bot_name == "search":
        headers = search_column_headers
    else:
        headers = scout_column_headers
    for header in headers:
        cell = sheet.cell(row=1, column=headers.index(header) + 1, value=(header.replace("_", " ")).title())
        cell.font = font
        cell.alignment  = Alignment(wrap_text=True)
        # For alternate coloring
        # if headers.index(header) % 2 == 0:
        #     cell.fill = PatternFill(
        #             fill_type="solid",
        #             fgColor='ff9b00'
        #           )
        # else:
        #     cell.fill = PatternFill(
        #             fill_type="solid",
        #             fgColor='0064d2'
        #           )
            # pass
    return wb
# wb.freeze_panes = "A2"
# wb.save("test.xlsx")            

# a = init_excel()
# a.save("test.xlsx")
